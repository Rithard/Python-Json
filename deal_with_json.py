#coding:utf-8
#Author:UINWZ

import json
import os
from openpyxl import Workbook

def get_filename():
    json_dir = os.path.split(os.path.realpath(__file__))[0]
    json_lists = list()
    for i in os.listdir(json_dir):
        if i.endswith('.json'):
            json_lists.append(i)
        else:
            pass
    return json_lists


def read_json(json_lists):
    for json_dir in json_lists:
        print(json_dir)
        with open(json_dir, 'r', encoding='utf-8') as f:
            info= list()
            for each_line in f.readlines():
                py_data = json.loads(each_line)
                cus_lists = py_data['result']['list']
                for v, cus_msg in enumerate(cus_lists):
                    info.append(cus_msg['ftm'])   #下单时间
                    info.append(cus_msg['oid'])   #订单编号
                    info.append(cus_msg['nam'])   #客户姓名
                    info.append(cus_msg['fad'])   #客户地址
                    info.append(cus_msg['olb'])   #订单状态
    return info


def create_excel(list,filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "订单信息"
    ws.cell(row=1, column=1).value = '订单编号'
    ws.cell(row=1, column=2).value = '下单时间'
    ws.cell(row=1, column=3).value = '客户姓名'
    ws.cell(row=1, column=4).value = '客户地址'
    ws.cell(row=1, column=5).value = '订单状态'
    rownum = 2

    i = 0
    while i< len(list):
        ws.cell(row=rownum, column=1).value = list[i]
        i += 1
        ws.cell(row=rownum, column=2).value = list[i]
        i += 1
        ws.cell(row=rownum, column=3).value = list[i]
        i += 1
        ws.cell(row=rownum, column=4).value = list[i]
        i += 1
        ws.cell(row=rownum, column=5).value = list[i]
        i += 1
        rownum += 1
    wb.save( filename + '.xlsx')
    print('Excel生成成功!')
if __name__ == '__main__':
    test_jingm = 'test_jingming'
    create_excel(read_json(get_filename()),test_jingm)
